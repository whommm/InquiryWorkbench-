"""
Excel export service for inquiry sheets
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from typing import List


def export_sheet_to_excel(sheet_data: List[List], filename: str = "inquiry_sheet.xlsx") -> BytesIO:
    """
    Export sheet data to Excel format with styling

    Args:
        sheet_data: 2D list of sheet data (first row is headers)
        filename: Output filename (not used, kept for compatibility)

    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "询价单"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write data
    for row_idx, row_data in enumerate(sheet_data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            cell.border = thin_border

            # Apply header styling to first row
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
